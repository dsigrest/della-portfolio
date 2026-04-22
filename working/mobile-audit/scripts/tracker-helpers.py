"""
openpyxl helpers for audit-tracker.xlsx.

Never use pandas on this file — pandas overwrites clobber concurrent edits. Always
read → modify specific rows → save via openpyxl with atomic patterns.

Usage:
    from tracker_helpers import read_tracker, append_row, update_row, init_tracker
"""
from pathlib import Path
from datetime import date

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    import sys
    print("Missing dependency. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)

COLUMNS = [
    "diagram_id",
    "case_study",
    "file_path",
    "figma_node_id",
    "ok_1440",
    "ok_1024",
    "ok_768",
    "ok_480",
    "ok_375",
    "ok_320",
    "severity",
    "root_cause",
    "fix_strategy",
    "mobile_file_path",
    "figma_mobile_node_id",
    "status",
    "audit_date",
    "verify_date",
    "notes",
]

VALID_STATUSES = {"pending", "audited", "fix_in_progress", "fixed", "verified", "blocked"}

def init_tracker(path):
    """Create a new tracker xlsx with headers. Safe to re-run — skips if exists."""
    p = Path(path)
    if p.exists():
        return p
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "audit"
    ws.append(COLUMNS)
    # style header
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    # freeze header row
    ws.freeze_panes = "A2"
    # column widths
    widths = {"A": 22, "B": 22, "C": 50, "D": 18, "K": 10, "L": 35, "M": 40, "N": 50, "O": 18, "P": 16, "Q": 12, "R": 12, "S": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    wb.save(p)
    return p

def validate_schema(path):
    """Confirm columns match expected schema. Raises if not."""
    wb = load_workbook(path, read_only=True)
    ws = wb["audit"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if header != COLUMNS:
        raise ValueError(f"Schema mismatch. Expected {COLUMNS}, got {header}")
    wb.close()
    return True

def read_tracker(path):
    """Return list of row dicts. Read-only, no lock held."""
    p = Path(path)
    if not p.exists():
        return []
    wb = load_workbook(p, read_only=True)
    ws = wb["audit"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(COLUMNS, r)))
    wb.close()
    return rows

def append_row(path, row_dict):
    """Append one row. Atomic: load → append → save → close."""
    _validate_row(row_dict)
    p = Path(path)
    if not p.exists():
        init_tracker(p)
    wb = load_workbook(p)
    ws = wb["audit"]
    values = [row_dict.get(c, "") for c in COLUMNS]
    ws.append(values)
    wb.save(p)
    wb.close()

def update_row(path, diagram_id, updates):
    """Find row by diagram_id and update specific columns. Atomic."""
    _validate_partial(updates)
    p = Path(path)
    wb = load_workbook(p)
    ws = wb["audit"]
    id_col = COLUMNS.index("diagram_id") + 1
    updated = False
    for r in ws.iter_rows(min_row=2):
        if r[id_col - 1].value == diagram_id:
            for col_name, new_val in updates.items():
                col_idx = COLUMNS.index(col_name)
                r[col_idx].value = new_val
            updated = True
            break
    if not updated:
        wb.close()
        raise KeyError(f"No row for diagram_id={diagram_id}")
    wb.save(p)
    wb.close()

def upsert_row(path, row_dict):
    """Append if diagram_id not present; else update."""
    _validate_row(row_dict)
    existing = {r["diagram_id"] for r in read_tracker(path)}
    if row_dict["diagram_id"] in existing:
        update_row(path, row_dict["diagram_id"], {k: v for k, v in row_dict.items() if k != "diagram_id"})
    else:
        append_row(path, row_dict)

def _validate_row(row_dict):
    if "diagram_id" not in row_dict:
        raise ValueError("row_dict missing required diagram_id")
    _validate_partial(row_dict)

def _validate_partial(updates):
    for k in updates:
        if k not in COLUMNS:
            raise ValueError(f"Unknown column: {k}. Valid: {COLUMNS}")
    if "status" in updates and updates["status"] not in VALID_STATUSES:
        raise ValueError(f"Invalid status: {updates['status']}. Valid: {VALID_STATUSES}")
    if "severity" in updates:
        sev = updates["severity"]
        if sev is not None and sev not in (0, 1, 2, 3):
            raise ValueError(f"Invalid severity: {sev}. Valid: 0, 1, 2, 3")

if __name__ == "__main__":
    # init a fresh tracker when run standalone
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "working/mobile-audit/audit-tracker.xlsx"
    p = init_tracker(path)
    validate_schema(p)
    print(f"Tracker initialized at {p}")
